import win32com.client
import os
from openpyxl import Workbook
from docx import Document

def check_outlook_inbox(limit=5):
    """Read top emails from Outlook inbox."""
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)  # Inbox
    messages = inbox.Items
    emails = []
    for i, message in enumerate(messages, start=1):
        if i > limit:
            break
        emails.append(f"From: {message.SenderName}, Subject: {message.Subject}")
    return emails

def create_excel(file_path="example.xlsx"):
    """Create a sample Excel file."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws['A1'] = "Hello"
    ws['B1'] = "World"
    wb.save(file_path)
    return f"Excel file created at {file_path}"

def create_word_doc(file_path="example.docx", text="Hello from AI Agent!"):
    """Create a sample Word document."""
    doc = Document()
    doc.add_paragraph(text)
    doc.save(file_path)
    return f"Word document created at {file_path}"

def open_app(app_name="notepad.exe"):
    """Open a local application."""
    try:
        # Handle special cases for common applications
        if app_name == "winword.exe" or app_name == "word":
            # Try multiple ways to open Word
            word_paths = [
                "winword.exe",  # If Word is in PATH
                "microsoft word",  # Windows start menu
                "word"  # Simple name
            ]
            
            for path in word_paths:
                try:
                    os.startfile(path)
                    return "Microsoft Word opened successfully"
                except OSError:
                    continue
            
            # If all methods fail
            return "Could not open Microsoft Word. Please make sure it's installed."
        
        # For other applications
        os.startfile(app_name)
        return f"Opened {app_name}"
        
    except Exception as e:
        return f"Error opening {app_name}: {e}"
